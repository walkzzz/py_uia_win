# 数据IO模块
# 提供多种数据格式的读写功能，包括文本文件、Excel文件、JSON文件和CSV文件

from typing import List, Dict, Any, Optional, Union
import os
import json
import csv
from datetime import datetime

# 尝试导入openpyxl用于Excel文件处理
try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.worksheet.worksheet import Worksheet
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

class DataIO:
    """数据IO类"""
    
    @staticmethod
    def read_text_file(file_path: str, encoding: str = "utf-8") -> str:
        """读取文本文件

        Args:
            file_path: 文件路径
            encoding: 文件编码

        Returns:
            文件内容

        Example:
            | ${content} | Read Text File | C:/test.txt | encoding=utf-8 |
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        with open(file_path, "r", encoding=encoding) as f:
            return f.read()

    @staticmethod
    def write_text_file(file_path: str, content: str, encoding: str = "utf-8", mode: str = "w") -> None:
        """写入文本文件

        Args:
            file_path: 文件路径
            content: 文件内容
            encoding: 文件编码
            mode: 写入模式（w: 覆盖, a: 追加）

        Example:
            | Write Text File | C:/test.txt | Hello World | encoding=utf-8 | mode=w |
            | Write Text File | C:/test.txt | \nAppend Line | encoding=utf-8 | mode=a |
        """
        # 创建目录（如果不存在）
        dir_path = os.path.dirname(file_path)
        if dir_path and not os.path.exists(dir_path):
            os.makedirs(dir_path)

        with open(file_path, mode, encoding=encoding) as f:
            f.write(content)

    @staticmethod
    def read_lines(file_path: str, encoding: str = "utf-8", strip: bool = False) -> List[str]:
        """按行读取文件

        Args:
            file_path: 文件路径
            encoding: 文件编码
            strip: 是否去除每行的前后空白字符

        Returns:
            行列表

        Example:
            | ${lines} | Read Lines | C:/test.txt | encoding=utf-8 | strip=True |
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        with open(file_path, "r", encoding=encoding) as f:
            lines = f.readlines()

        if strip:
            lines = [line.strip() for line in lines]
        else:
            lines = [line.rstrip('\n') for line in lines]

        return lines

    @staticmethod
    def write_lines(file_path: str, lines: List[str], encoding: str = "utf-8", mode: str = "w") -> None:
        """按行写入文件

        Args:
            file_path: 文件路径
            lines: 行列表
            encoding: 文件编码
            mode: 写入模式（w: 覆盖, a: 追加）

        Example:
            | ${lines} | Create List | Line 1 | Line 2 | Line 3 |
            | Write Lines | C:/test.txt | ${lines} | encoding=utf-8 | mode=w |
        """
        # 创建目录（如果不存在）
        dir_path = os.path.dirname(file_path)
        if dir_path and not os.path.exists(dir_path):
            os.makedirs(dir_path)

        with open(file_path, mode, encoding=encoding) as f:
            for line in lines:
                f.write(f"{line}\n")

    @staticmethod
    def read_json(file_path: str, encoding: str = "utf-8") -> Dict[str, Any]:
        """读取JSON文件

        Args:
            file_path: 文件路径
            encoding: 文件编码

        Returns:
            JSON数据

        Example:
            | ${data} | Read Json | C:/test.json | encoding=utf-8 |
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        with open(file_path, "r", encoding=encoding) as f:
            return json.load(f)

    @staticmethod
    def write_json(file_path: str, data: Dict[str, Any], encoding: str = "utf-8", indent: int = 4) -> None:
        """写入JSON文件

        Args:
            file_path: 文件路径
            data: JSON数据
            encoding: 文件编码
            indent: 缩进空格数

        Example:
            | ${data} | Create Dictionary | name=test | value=123 |
            | Write Json | C:/test.json | ${data} | encoding=utf-8 | indent=4 |
        """
        # 创建目录（如果不存在）
        dir_path = os.path.dirname(file_path)
        if dir_path and not os.path.exists(dir_path):
            os.makedirs(dir_path)

        with open(file_path, "w", encoding=encoding) as f:
            json.dump(data, f, ensure_ascii=False, indent=indent)

    @staticmethod
    def read_csv(file_path: str, encoding: str = "utf-8", delimiter: str = ",", header: bool = True) -> Union[List[Dict[str, Any]], List[List[str]]]:
        """读取CSV文件

        Args:
            file_path: 文件路径
            encoding: 文件编码
            delimiter: 分隔符
            header: 是否包含表头

        Returns:
            CSV数据
            - 如果header=True，返回字典列表 [{"列1": "值1", "列2": "值2"}]
            - 如果header=False，返回列表列表 [["值1", "值2"]]

        Example:
            | ${data} | Read Csv | C:/test.csv | encoding=utf-8 | delimiter=, | header=True |
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        with open(file_path, "r", encoding=encoding) as f:
            reader = csv.reader(f, delimiter=delimiter)
            rows: List[List[str]] = list(reader)

        if not rows:
            return []

        if header:
            # 第一行为表头
            headers = rows[0]
            data = []
            for row in rows[1:]:
                # 处理行长度与表头长度不匹配的情况
                row_dict = {}
                for i, header in enumerate(headers):
                    if i < len(row):
                        row_dict[header] = row[i]
                    else:
                        row_dict[header] = ""
                data.append(row_dict)
            return data
        else:
            return rows

    @staticmethod
    def write_csv(file_path: str, data: Union[List[Dict[str, Any]], List[List[str]]], encoding: str = "utf-8", delimiter: str = ",", header: Optional[List[str]] = None) -> None:
        """写入CSV文件

        Args:
            file_path: 文件路径
            data: CSV数据
            encoding: 文件编码
            delimiter: 分隔符
            header: 表头列表，如果为None则自动从数据中提取

        Example:
            | ${data} | Create List | Create Dictionary | name=test1 | value=123 | Create Dictionary | name=test2 | value=456 |
            | Write Csv | C:/test.csv | ${data} | encoding=utf-8 | delimiter=, |
        """
        # 创建目录（如果不存在）
        dir_path = os.path.dirname(file_path)
        if dir_path and not os.path.exists(dir_path):
            os.makedirs(dir_path)

        with open(file_path, "w", newline="", encoding=encoding) as f:
            writer = csv.writer(f, delimiter=delimiter)

            if isinstance(data, list) and data and isinstance(data[0], dict):
                # 数据是字典列表
                if header is None:
                    # 自动提取表头
                    header = list(data[0].keys())  # type: ignore[assignment]

                # 写入表头
                writer.writerow(header)  # type: ignore[arg-type]

                # 写入数据
                for row in data:
                    writer.writerow([row.get(h, "") for h in header])  # type: ignore[arg-type]
            else:
                # 数据是列表列表
                if header:
                    writer.writerow(header)

                for row in data:
                    writer.writerow(row)

    @staticmethod
    def read_excel(file_path: str, sheet_name: Optional[str] = None, header: bool = True, start_row: int = 1) -> Union[List[Dict[str, Any]], List[List[str]]]:
        """读取Excel文件

        Args:
            file_path: 文件路径
            sheet_name: 工作表名称，如果为None则读取第一个工作表
            header: 是否包含表头
            start_row: 开始行号（从1开始）

        Returns:
            Excel数据
            - 如果header=True，返回字典列表 [{"列1": "值1", "列2": "值2"}]
            - 如果header=False，返回列表列表 [["值1", "值2"]]

        Example:
            | ${data} | Read Excel | C:/test.xlsx | sheet_name=Sheet1 | header=True | start_row=1 |
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is not installed. Please install it with 'pip install openpyxl'")

        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        # 加载工作簿
        wb = load_workbook(file_path, data_only=True)

        # 选择工作表
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        # 读取数据
        data = []
        for row in ws.iter_rows(min_row=start_row, values_only=True):
            # 过滤空行
            if any(cell is not None and cell != "" for cell in row):
                data.append(list(row))

        if not data:
            return []

        if header:
            # 第一行为表头
            headers = data[0]
            rows = []
            for row in data[1:]:
                # 处理行长度与表头长度不匹配的情况
                row_dict = {}
                for i, header in enumerate(headers):
                    if i < len(row) and row[i] is not None:
                        row_dict[header] = row[i]
                    else:
                        row_dict[header] = ""
                rows.append(row_dict)
            return rows
        else:
            return data

    @staticmethod
    def write_excel(file_path: str, data: Union[List[Dict[str, Any]], List[List[str]]], sheet_name: str = "Sheet1", header: Optional[List[str]] = None) -> None:
        """写入Excel文件

        Args:
            file_path: 文件路径
            data: Excel数据
            sheet_name: 工作表名称
            header: 表头列表，如果为None则自动从数据中提取

        Example:
            | ${data} | Create List | Create Dictionary | name=test1 | value=123 | Create Dictionary | name=test2 | value=456 |
            | Write Excel | C:/test.xlsx | ${data} | sheet_name=Sheet1 |
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is not installed. Please install it with 'pip install openpyxl'")

        # 创建工作簿
        wb = Workbook()

        # 创建或选择工作表
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
            # 删除默认工作表
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        # 写入数据
        if isinstance(data, list) and data and isinstance(data[0], dict):
            # 数据是字典列表
            if header is None:
                # 自动提取表头
                header = list(data[0].keys())

            # 写入表头
            ws.append(header)

            # 写入数据
            for row in data:
                ws.append([row.get(h, "") for h in header])
        else:
            # 数据是列表列表
            if header:
                ws.append(header)

            for row in data:
                ws.append(row)

        # 保存工作簿
        wb.save(file_path)

    @staticmethod
    def append_to_excel(file_path: str, data: Union[List[Dict[str, Any]], List[List[str]]], sheet_name: str = "Sheet1") -> None:
        """追加数据到Excel文件

        Args:
            file_path: 文件路径
            data: Excel数据
            sheet_name: 工作表名称

        Example:
            | ${data} | Create List | Create Dictionary | name=test3 | value=789 |
            | Append To Excel | C:/test.xlsx | ${data} | sheet_name=Sheet1 |
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is not installed. Please install it with 'pip install openpyxl'")

        if not os.path.exists(file_path):
            # 文件不存在，直接创建
            DataIO.write_excel(file_path, data, sheet_name)
            return

        # 加载工作簿
        wb = load_workbook(file_path)

        # 选择工作表
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)

        # 追加数据
        if isinstance(data, list) and data and isinstance(data[0], dict):
            # 数据是字典列表
            # 获取现有表头
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)

            # 写入数据
            for row in data:
                ws.append([row.get(h, "") for h in headers])
        else:
            # 数据是列表列表
            for row in data:
                ws.append(row)

        # 保存工作簿
        wb.save(file_path)

    @staticmethod
    def get_file_lines_count(file_path: str, encoding: str = "utf-8") -> int:
        """获取文件行数

        Args:
            file_path: 文件路径
            encoding: 文件编码

        Returns:
            文件行数

        Example:
            | ${lines_count} | Get File Lines Count | C:/test.txt | encoding=utf-8 |
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        with open(file_path, "r", encoding=encoding) as f:
            return sum(1 for _ in f)

    @staticmethod
    def get_file_size(file_path: str) -> int:
        """获取文件大小（字节）

        Args:
            file_path: 文件路径

        Returns:
            文件大小（字节）

        Example:
            | ${size} | Get File Size | C:/test.txt |
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        return os.path.getsize(file_path)

    @staticmethod
    def get_file_mtime(file_path: str) -> str:
        """获取文件修改时间

        Args:
            file_path: 文件路径

        Returns:
            文件修改时间，格式：YYYY-MM-DD HH:MM:SS

        Example:
            | ${mtime} | Get File Mtime | C:/test.txt |
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        mtime = os.path.getmtime(file_path)
        return datetime.fromtimestamp(mtime).strftime("%Y-%m-%d %H:%M:%S")

    @staticmethod
    def copy_file(src: str, dst: str) -> None:
        """复制文件

        Args:
            src: 源文件路径
            dst: 目标文件路径

        Example:
            | Copy File | C:/source.txt | C:/destination.txt |
        """
        if not os.path.exists(src):
            raise FileNotFoundError(f"Source file not found: {src}")

        # 创建目标目录
        dst_dir = os.path.dirname(dst)
        if dst_dir and not os.path.exists(dst_dir):
            os.makedirs(dst_dir)

        import shutil
        shutil.copy2(src, dst)

    @staticmethod
    def move_file(src: str, dst: str) -> None:
        """移动文件

        Args:
            src: 源文件路径
            dst: 目标文件路径

        Example:
            | Move File | C:/source.txt | C:/destination.txt |
        """
        if not os.path.exists(src):
            raise FileNotFoundError(f"Source file not found: {src}")

        # 创建目标目录
        dst_dir = os.path.dirname(dst)
        if dst_dir and not os.path.exists(dst_dir):
            os.makedirs(dst_dir)

        import shutil
        shutil.move(src, dst)

    @staticmethod
    def delete_file(file_path: str) -> None:
        """删除文件

        Args:
            file_path: 文件路径

        Example:
            | Delete File | C:/test.txt |
        """
        if os.path.exists(file_path):
            os.remove(file_path)

    @staticmethod
    def file_exists(file_path: str) -> bool:
        """检查文件是否存在

        Args:
            file_path: 文件路径

        Returns:
            文件是否存在

        Example:
            | ${exists} | File Exists | C:/test.txt |
        """
        return os.path.exists(file_path)

    @staticmethod
    def create_directory(dir_path: str) -> None:
        """创建目录

        Args:
            dir_path: 目录路径

        Example:
            | Create Directory | C:/new_dir |
        """
        if not os.path.exists(dir_path):
            os.makedirs(dir_path)

    @staticmethod
    def delete_directory(dir_path: str, recursive: bool = False) -> None:
        """删除目录

        Args:
            dir_path: 目录路径
            recursive: 是否递归删除

        Example:
            | Delete Directory | C:/empty_dir |
            | Delete Directory | C:/non_empty_dir | recursive=True |
        """
        if os.path.exists(dir_path):
            if recursive:
                import shutil
                shutil.rmtree(dir_path)
            else:
                os.rmdir(dir_path)

    @staticmethod
    def list_files(dir_path: str, pattern: Optional[str] = None) -> List[str]:
        """列出目录中的文件

        Args:
            dir_path: 目录路径
            pattern: 文件名称模式

        Returns:
            文件路径列表

        Example:
            | ${files} | List Files | C:/test_dir | pattern=*.txt |
        """
        if not os.path.exists(dir_path):
            raise FileNotFoundError(f"Directory not found: {dir_path}")

        files = []
        for file in os.listdir(dir_path):
            file_path = os.path.join(dir_path, file)
            if os.path.isfile(file_path):
                if pattern is None or file.endswith(pattern) or file == pattern:
                    files.append(file_path)

        return files

    @staticmethod
    def list_directories(dir_path: str) -> List[str]:
        """列出目录中的子目录

        Args:
            dir_path: 目录路径

        Returns:
            子目录路径列表

        Example:
            | ${dirs} | List Directories | C:/parent_dir |
        """
        if not os.path.exists(dir_path):
            raise FileNotFoundError(f"Directory not found: {dir_path}")

        dirs = []
        for dir_name in os.listdir(dir_path):
            dir_path_full = os.path.join(dir_path, dir_name)
            if os.path.isdir(dir_path_full):
                dirs.append(dir_path_full)

        return dirs

# 创建DataIO实例
data_io = DataIO()
